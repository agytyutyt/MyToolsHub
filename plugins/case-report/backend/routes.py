"""战果录入 —— JZToolsHub 后端插件路由。

功能：输入公安某部门对某案件的「收网情况报告」，抽取五要素
（案件名 / 时间 / 主办大队 / 抓获人数 / 缴获物品），
以「键值对」JSON 的形式本地化存档（backend/data/ 下，一记录一文件）。

缴获物品：除保留原文摘要外，逐项拆分为 {category, name, quantity, unit}
单列存储；类似物品（如电脑/笔记本）归为统一战果类别，可跨记录按
「战果汇总」（/aggregate）做数量叠加。

解析策略：仅大模型解析。未配置大模型时返回错误提示，前端将会提示用户配置。
（已取消本地规则与正则解析，所有字段与缴获明细均由大模型结构化输出。）

时间规整：仅写「月/日」时自动补当前年份；出现「昨天/前天/今天」等
以当前年月日为基准回推。

并发设计：解析含大模型调用，走后台线程池 + task_id 轮询，
不占用 HTTP worker（与 character-graph 一致）。

接口前缀：/api/case-report
"""

import io
import json
import os
import re
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime

from flask import jsonify, request, send_file

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    openpyxl = None
    OPENPYXL_AVAILABLE = False

from . import category_kb, llm_client, parser

try:
    from jztools_admin.routes import get_session_user as _get_session_user
except Exception:  # admin 插件缺失时兜底（理论上不会发生）
    _get_session_user = None

try:
    from jztools_admin.routes import set_operation as _set_operation
except Exception:  # 主应用未提供日志辅助时兜底（理论上不会发生）
    def _set_operation(op):
        pass

try:
    import requests  # noqa: F401  大模型 HTTP 调用依赖
    REQUESTS_AVAILABLE = True
except Exception:
    requests = None
    REQUESTS_AVAILABLE = False

import jztools_data
CONFIG_FILE = jztools_data.get_data_root_file("plugins", "case-report", "config.json")
PROMPT_FILE = jztools_data.get_data_root_file("plugins", "case-report", "prompt.json")
DATA_DIR = jztools_data.get_data_root_dir("plugins", "case-report", "data")
# 主办大队可选值配置文件（插件目录内固化，随插件分发）
ORG_UNITS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "org_units.json")
API_PREFIX = "/api/case-report"

# 后台解析线程池：限制并发大模型任务数
PARSE_WORKERS = 2
_parse_executor = ThreadPoolExecutor(max_workers=PARSE_WORKERS)

# 任务状态：task_id -> {status, ...}; pending -> running -> done / error
TASKS = {}
TASKS_LOCK = threading.Lock()
TASK_TTL_SECONDS = 30 * 60

# 全局文件写锁（记录读写、配置写共用）
_LOCK = threading.RLock()

RECORD_ID_RE = re.compile(r"^[A-Za-z0-9_-]+$")
MAX_TEXT_LEN = 20000
MAX_ITEM_LEN = 2000


def _case_sort_key(name):
    """中文案件名的拼音排序键（零依赖）。

    GB2312/GBK 一级常用汉字按拼音排序，直接比较其 GBK 编码即可近似拼音序；
    数字/字母/标点（ASCII）排在中文前，无法用 GBK 编码的生僻字排最后。
    """
    n = (name or "").strip()
    if not n:
        return (0, b"")
    try:
        return (0, n.encode("gbk"))
    except UnicodeEncodeError:
        return (1, n)

DEFAULT_CONFIG = {"llm": {"base_url": "", "api_key": "", "model": ""}}
FIELD_KEYS = parser.FIELD_KEYS


# ===================== 基础 IO =====================

def _now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def load_config():
    data = {}
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = {}
    llm = data.get("llm") or {}
    return {"llm": {
        "base_url": (llm.get("base_url") or "").strip(),
        "api_key": (llm.get("api_key") or "").strip(),
        "model": (llm.get("model") or "").strip(),
    }}


def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def load_prompt():
    data = {}
    if os.path.exists(PROMPT_FILE):
        try:
            with open(PROMPT_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = {}
    return {
        "system": data.get("system") or llm_client.DEFAULT_SYSTEM_PROMPT,
        "user_template": data.get("user_template") or llm_client.DEFAULT_USER_TEMPLATE,
    }


def load_org_units():
    """读取主办大队可选值（插件目录 org_units.json 固化）。"""
    units = []
    try:
        with open(ORG_UNITS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        units = [str(u).strip() for u in (data.get("主办大队") or []) if str(u).strip()]
    except Exception:
        pass
    return units


def ensure_files():
    """首次运行生成可编辑的默认配置文件与数据目录。"""
    if not os.path.exists(CONFIG_FILE):
        save_config(DEFAULT_CONFIG)
    if not os.path.exists(PROMPT_FILE):
        with open(PROMPT_FILE, "w", encoding="utf-8") as f:
            json.dump({
                "system": llm_client.DEFAULT_SYSTEM_PROMPT,
                "user_template": llm_client.DEFAULT_USER_TEMPLATE,
            }, f, ensure_ascii=False, indent=2)
    os.makedirs(DATA_DIR, exist_ok=True)


# ===================== 字段规整 =====================

def _normalize_unit(dept, units=None):
    """把主办大队归一到配置文件可选值（一大队/二大队/三大队）。

    规则：已是可选值之一则保留；否则从文本中提取中文/阿拉伯数字 1-3，
    映射到对应序号的可选值（一大队→1、二大队→2、三大队→3）；
    无法识别则置空。
    """
    units = units if units is not None else load_org_units()
    if not units:
        return (dept or "").strip()
    dept = (dept or "").strip()
    if dept in units:
        return dept
    if not dept:
        return ""
    cn = {"一": 1, "二": 2, "三": 3}
    num = None
    m = re.search(r"([一二三1-3])", dept)
    if m:
        ch = m.group(1)
        num = cn.get(ch) if ch in cn else int(ch)
    if num and 1 <= num <= len(units):
        return units[num - 1]
    return ""


def normalize_fields(raw, now=None):
    """只保留五要素，trim 长度，抓获人数转阿拉伯数字/整数，时间补全整日期。

    支持以下自动补充：
    - 时间字段为空时填入当前日期（YYYY年M月D日）；
    - 主办大队归一到插件配置的可选值（一大队/二大队/三大队）。
    """
    now = now or datetime.now()
    raw = raw or {}
    out = {}
    for k in FIELD_KEYS:
        v = raw.get(k)
        if not isinstance(v, (str, int, float)):
            v = ""
        v = str(v).strip()
        if len(v) > MAX_ITEM_LEN:
            v = v[:MAX_ITEM_LEN]
        out[k] = v
    # 时间：为空时填入当前日期
    if out["时间"]:
        out["时间"] = parser.normalize_time(out["时间"])
    if not out["时间"]:
        out["时间"] = f"{now.year}年{now.month}月{now.day}日"
    # 主办大队：归一到配置可选值
    out["主办大队"] = _normalize_unit(out["主办大队"])
    cnt = out["抓获人数"]
    if cnt:
        cnt = parser.normalize_count(cnt)
        out["抓获人数"] = cnt
        if cnt.isdigit():
            out["抓获人数"] = int(cnt)
        else:
            out["抓获人数"] = cnt
    return out


# ===================== 解析任务（后台线程池） =====================

def set_task(task_id, **kwargs):
    with TASKS_LOCK:
        TASKS[task_id] = {**TASKS.get(task_id, {}), **kwargs}


def get_task(task_id):
    with TASKS_LOCK:
        return TASKS.get(task_id)


def cleanup_tasks():
    now = time.time()
    with TASKS_LOCK:
        expired = [tid for tid, t in TASKS.items()
                   if t.get("created_at", 0) < now - TASK_TTL_SECONDS]
        for tid in expired:
            TASKS.pop(tid, None)


def create_task(created_at):
    tid = uuid.uuid4().hex[:12]
    set_task(tid, status="pending", created_at=created_at)
    return tid


def _build_items(llm_items):
    """把大模型返回的缴获物品明细数组规整为 items（含用户学习类别覆盖）。

    llm_items: [{"名称","类别","数量","单位"}, ...]。
    """
    items = []
    for it in llm_items or []:
        if not isinstance(it, dict):
            continue
        name = str(it.get("名称") or "").strip()
        if not name:
            continue
        qty = it.get("数量")
        try:
            qty = round(float(qty), 2) if qty not in (None, "") else None
        except (TypeError, ValueError):
            qty = None
        items.append({
            "category": str(it.get("类别") or "").strip() or parser.CATEGORY_OTHER,
            "name": name,
            "quantity": qty,
            "unit": str(it.get("单位") or "").strip(),
        })
    overrides = category_kb.all_overrides()
    return parser.apply_category_overrides(items, overrides)


def _run_parse(task_id, text, cfg_llm, api_key, model, prompt):
    """后台线程执行：仅大模型解析（未配置/失败时任务置 error，前端提示）。"""
    try:
        set_task(task_id, status="running")
        llm_error = ""
        api_key = (api_key or "").strip()
        if not api_key:
            llm_error = "未配置大模型解析：请先在「大模型解析配置」中填写 API 地址、API Key 与模型名称"
            set_task(task_id, status="error", detail=llm_error)
            return

        try:
            llm_fields = llm_client.extract_fields(
                text, cfg_llm.get("base_url"), api_key,
                model or cfg_llm.get("model"), prompt,
                timeout=120,
            )
        except llm_client.LLMError as e:
            llm_error = str(e)
            set_task(task_id, status="error", detail=f"大模型解析失败：{llm_error}")
            return
        except Exception as e:
            llm_error = f"大模型调用异常（{type(e).__name__}）"
            set_task(task_id, status="error", detail=llm_error)
            return

        fields = normalize_fields(llm_fields)
        fields.pop("主办人", None)  # 主办人不参与解析，留空由保存时默认用户姓名填充
        items = _build_items(llm_fields.get("缴获物品明细"))
        set_task(task_id, status="done", fields=fields, method="llm",
                 llm_error="", items=items)
    except (ValueError, RuntimeError) as e:
        set_task(task_id, status="error", detail=str(e))
    except Exception as e:
        # SEC-5：非预期异常不向前端透出内部细节（堆栈/路径等）
        set_task(task_id, status="error", detail=f"解析失败（{type(e).__name__}）")


def _submit_parse(text, cfg_llm, api_key, model, prompt):
    task_id = create_task(time.time())
    cleanup_tasks()
    _parse_executor.submit(_run_parse, task_id, text, cfg_llm, api_key, model, prompt)
    return task_id


# ===================== 记录存储（backend/data/，一记录一 JSON 文件） =====================

def _record_path(rid):
    return os.path.join(DATA_DIR, f"{rid}.json")


def _check_rid(rid):
    return bool(RECORD_ID_RE.match(rid or ""))


def _is_record(rec):
    """记录必须含 id/fields（用于区分台账文件与 item_categories.json 等非记录文件）。"""
    return isinstance(rec, dict) and bool(rec.get("id")) and isinstance(rec.get("fields"), dict)


def load_record(rid):
    if not _check_rid(rid):
        return None
    path = _record_path(rid)
    if not os.path.isfile(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            rec = json.load(f)
        if not _is_record(rec):
            return None
        return rec
    except Exception:
        return None


def save_record(rec):
    """原子写：先写临时文件再替换，避免半截 JSON 落盘。"""
    os.makedirs(DATA_DIR, exist_ok=True)
    path = _record_path(rec["id"])
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(rec, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def _find_case_matches(case_name, recs):
    """在指定记录集中查找与规整案件名相同的既有记录（用于同案合并提示）。

    recs 由调用方传入已按归属过滤的记录集：默认传 scoped_records(user, "mine")，
    保证同案合并目标始终是自己可管理的记录，不会把战果并入他人案件。
    """
    out = []
    for rec in recs:
        rn = parser.normalize_case_name((rec.get("fields") or {}).get("案件名"))
        if rn and rn == case_name:
            out.append({
                "id": rec["id"],
                "name": (rec.get("fields") or {}).get("案件名") or "",
                "created_at": rec.get("created_at") or "",
                "items": len(rec.get("items") or []),
            })
    out.sort(key=lambda r: r.get("created_at") or "", reverse=True)
    return out


def _validate_items(raw):
    """校验并规整前端交回的单列物品明细（含用户修正后的类别）。"""
    if not isinstance(raw, list):
        return None
    out = []
    for it in raw:
        if not isinstance(it, dict):
            continue
        name = str(it.get("name") or "").strip()
        if not name:
            continue
        category = str(it.get("category") or "").strip() or "其他"
        unit = str(it.get("unit") or "").strip()
        q = it.get("quantity")
        qty = None
        if q is not None and q not in ("", []):
            try:
                qty = round(float(q), 2)
            except (TypeError, ValueError):
                qty = None
        out.append({
            "category": category[:20],
            "name": name[:40],
            "quantity": qty,
            "unit": unit[:10],
        })
    return out


def list_records():
    """列出全部记录，按入库时间倒序。"""
    os.makedirs(DATA_DIR, exist_ok=True)
    recs = []
    for name in os.listdir(DATA_DIR):
        if not name.endswith(".json") or name.endswith(".tmp.json"):
            continue
        rid = name[:-5]
        rec = load_record(rid)
        if rec:
            recs.append(rec)
    recs.sort(key=lambda r: r.get("created_at", ""), reverse=True)
    return recs


# ===================== 数据归属与可见性 =====================

def _cr_viewer():
    """当前登录用户上下文；未登录返回 None。"""
    if _get_session_user is None:
        return None
    try:
        return _get_session_user()
    except Exception:
        return None


def scoped_records(user, scope):
    """按查看范围过滤台账记录。

    scope: mine=我的(默认)；dept=同部门；all=全站(仅超管可用)
    """
    recs = list_records()
    if scope == "all" and user.get("super_admin"):
        return recs
    if scope == "dept":
        return [
            r for r in recs
            if r.get("created_by") == user["username"]
            or (r.get("department_id") == user.get("department_id")
                and r.get("unit_id") == user.get("unit_id"))
        ]
    return [r for r in recs if r.get("created_by") == user["username"]]


def _record_readable(user, rec):
    """读权限：超管 / 本人 / 同部门（部门+单位双重比对）。"""
    if user.get("super_admin"):
        return True
    if rec.get("created_by") == user["username"]:
        return True
    return (rec.get("department_id") == user.get("department_id")
            and rec.get("unit_id") == user.get("unit_id"))


def _record_manageable(user, rec):
    """写权限（改案件名 / 删除）：仅本人或超管。"""
    if user.get("super_admin"):
        return True
    return rec.get("created_by") == user["username"]


def _filter_by_period(recs, month, from_, to):
    """按入库时间（created_at）过滤记录。

    month=YYYY-MM 匹配该月；from / to 为 YYYY-MM-DD（或 YYYY-MM），按自然序比较。
    """
    out = recs
    if month:
        m = str(month)[:7]
        out = [r for r in out if (r.get("created_at") or "").startswith(m)]
    if from_:
        f = str(from_)[:10]
        out = [r for r in out if (r.get("created_at") or "")[:10] >= f]
    if to:
        t = str(to)[:10]
        out = [r for r in out if (r.get("created_at") or "")[:10] <= t]
    return out


def _collect_months(recs):
    """提取记录集中出现的「年-月」集合（去重、倒序），供前端月份下拉。"""
    seen = {}
    for r in recs:
        m = (r.get("created_at") or "")[:7]
        if m:
            seen[m] = True
    return sorted(seen.keys(), reverse=True)


def _filter_by_org(recs, department_id, username):
    """按部门 ID / 用户名（录入人）过滤记录；参数为空则不过滤。

    department_id：精确匹配记录 department_id（支持传入部门名称模糊匹配）；
    username：匹配 created_by 或 created_by_name（支持输入姓名/登录名搜索）。
    """
    out = recs
    dept = (department_id or "").strip()
    if dept:
        out = [r for r in out
               if (r.get("department_id") or "") == dept
               or dept in (r.get("department_name") or "")]
    user = (username or "").strip()
    if user:
        out = [r for r in out
               if (r.get("created_by") or "") == user
               or user in (r.get("created_by_name") or "")
               or (r.get("fields") or {}).get("主办人") == user]
    return out


def migrate_record_owner():
    """给历史记录补归属：录入人 admin。

    历史 records 的 unit/department 留空：admin 是超管走 super_admin 分支可见，
    不依赖这两个字段；空值也不会意外匹配到任何人的 department_id。
    """
    changed = False
    for rec in list_records():
        if rec.get("created_by"):
            continue
        rec["created_by"] = "admin"
        rec["created_by_name"] = "系统管理员"
        rec["unit_id"] = ""
        rec["department_id"] = ""
        save_record(rec)
        changed = True
    return changed


# ===================== 路由注册 =====================

def register(app) -> None:
    """插件入口：由 JZToolsHub 主应用在启动时调用。"""
    ensure_files()
    # 存量记录归属迁移（一次性、幂等）：给历史记录补默认录入人 admin
    with _LOCK:
        migrate_record_owner()

    @app.get(f"{API_PREFIX}/config")
    def cr_get_config():
        """读取展示配置。API Key 不回传明文，仅返回掩码（SEC-2/F-4）。"""
        cfg = load_config()
        key = cfg["llm"]["api_key"]
        masked = (key[:3] + "****" + key[-4:]) if len(key) > 8 else ("已设置" if key else "")
        return jsonify({
            "ok": True,
            "base_url": cfg["llm"]["base_url"],
            "api_key_set": bool(key),
            "api_key_masked": masked,
            "model": cfg["llm"]["model"],
            "llm_configured": bool(key) and bool(cfg["llm"]["base_url"]),
        })

    @app.get(f"{API_PREFIX}/org-units")
    def cr_get_org_units():
        """主办大队可选值（插件目录 org_units.json 固化）。"""
        return jsonify({"ok": True, "units": load_org_units()})

    @app.post(f"{API_PREFIX}/config")
    def cr_post_config():
        """保存展示配置；api_key 留空表示沿用原值，避免明文回显后空写覆盖。"""
        _set_operation("保存战果录入配置")
        body = request.get_json(silent=True) or {}
        merged = load_config()
        merged["llm"]["base_url"] = (body.get("base_url") or "").strip()
        new_key = (body.get("api_key") or "").strip()
        if new_key:
            merged["llm"]["api_key"] = new_key
        merged["llm"]["model"] = (body.get("model") or "").strip()
        save_config(merged)
        return jsonify({"ok": True, "llm_configured": bool(merged["llm"]["api_key"])})

    @app.post(f"{API_PREFIX}/config/test")
    def cr_test_config():
        """大模型连通性测试：用当前配置（或本次表单值）发一条最小请求验证。

        body 可选 {base_url?, api_key?, model?}，缺省回退到已保存配置；
        api_key 留空表示沿用已保存值（不回传明文的安全约定）。
        """
        _set_operation("测试大模型连通性")
        body = request.get_json(silent=True) or {}
        cfg = load_config()
        base_url = (body.get("base_url") or cfg["llm"]["base_url"] or "").strip()
        api_key = (body.get("api_key") or cfg["llm"]["api_key"] or "").strip()
        model = (body.get("model") or cfg["llm"]["model"] or "").strip()
        if not base_url:
            return jsonify({"ok": False, "detail": "请先填写 API 地址"}), 400
        if not api_key:
            return jsonify({"ok": False, "detail": "请先填写 API Key（或已保存配置）"}), 400
        ok, detail = llm_client.test_connection(base_url, api_key, model)
        return jsonify({"ok": ok, "detail": detail})

    @app.get(f"{API_PREFIX}/status")
    def cr_status():
        """依赖自检：缺依赖时优雅降级（B-4）。"""
        deps = {"requests": REQUESTS_AVAILABLE}
        return jsonify({"ok": all(deps.values()), "dependencies": deps})

    @app.post(f"{API_PREFIX}/parse")
    def cr_parse():
        """提交解析任务：body {text, base_url?, api_key?, model?}，立即返回 task_id。"""
        _set_operation("解析收网情况报告")
        body = request.get_json(silent=True) or {}
        text = body.get("text")
        if not text or not str(text).strip():
            return jsonify({"ok": False, "detail": "请先输入收网情况报告"}), 400
        text = str(text).strip()
        if len(text) > MAX_TEXT_LEN:
            return jsonify({"ok": False, "detail": f"报告过长，请控制在 {MAX_TEXT_LEN} 字以内"}), 400

        cfg = load_config()
        api_key = (body.get("api_key") or cfg["llm"]["api_key"] or "").strip()
        base_url = (body.get("base_url") or cfg["llm"]["base_url"] or "").strip()
        model = (body.get("model") or cfg["llm"]["model"] or "").strip()
        cfg_llm = {"base_url": base_url, "api_key": api_key, "model": model}

        task_id = _submit_parse(text, cfg_llm, api_key, model, load_prompt())
        return jsonify({"ok": True, "task_id": task_id})

    @app.get(f"{API_PREFIX}/result/<task_id>")
    def cr_result(task_id):
        """轮询解析状态：pending / running / done / error"""
        task = get_task(task_id)
        if not task:
            return jsonify({"ok": False, "detail": "任务不存在或已过期"}), 404
        payload = {"status": task.get("status")}
        if task.get("status") == "done":
            payload.update({
                "fields": task.get("fields"),
                "method": task.get("method", "llm"),
                "llm_error": task.get("llm_error", ""),
                "items": task.get("items", []),
            })
        if task.get("status") == "error":
            payload["detail"] = task.get("detail", "解析失败")
        return jsonify(payload)

    @app.get(f"{API_PREFIX}/aggregate")
    def cr_aggregate():
        """跨记录战果汇总：类似物品归为统一类别、数量叠加；
        「涉及 N 起」按规整后的案件名去重（同一案件多条记录只计一起）。
        支持 ?case=<案件名> 仅统计该案件的记录；支持 ?scope=mine|dept|all（默认 mine）按范围过滤；
        支持 ?month=YYYY-MM 或 ?from=...&to=... 按入库时间过滤。"""
        user = _cr_viewer()
        if user is None:
            return jsonify({"ok": False, "detail": "未登录或登录已过期"}), 401
        scope = (request.args.get("scope") or "mine").strip()
        if scope == "all" and not user.get("super_admin"):
            return jsonify({"ok": False, "detail": "无权查看全部战果"}), 403
        case_name = parser.normalize_case_name(request.args.get("case") or "")
        recs = scoped_records(user, scope)
        recs = _filter_by_period(recs,
                                 request.args.get("month") or "",
                                 request.args.get("from") or "",
                                 request.args.get("to") or "")
        recs = _filter_by_org(recs,
                              request.args.get("dept") or "",
                              request.args.get("user") or "")
        if case_name:
            recs = [r for r in recs
                    if parser.normalize_case_name((r.get("fields") or {}).get("案件名")) == case_name]
            # 单案视角：同案多条记录视为同一案件（涉及 N 起 = 1）
            case_keys = [case_name] * len(recs)
        else:
            case_keys = [
                parser.normalize_case_name((r.get("fields") or {}).get("案件名")) or r.get("id")
                for r in recs
            ]
        rows = parser.aggregate_items([r.get("items") or [] for r in recs], case_keys)
        return jsonify({"ok": True, "count": len(rows), "categories": rows})

    @app.get(f"{API_PREFIX}/months")
    def cr_months():
        """当前可见范围内的入库月份列表（倒序），供「按月份显示/统计」下拉使用。
        跟随 ?scope=mine|dept|all 过滤。"""
        user = _cr_viewer()
        if user is None:
            return jsonify({"ok": False, "detail": "未登录或登录已过期"}), 401
        scope = (request.args.get("scope") or "mine").strip()
        if scope == "all" and not user.get("super_admin"):
            return jsonify({"ok": False, "detail": "无权查看全部战果"}), 403
        months = _collect_months(scoped_records(user, scope))
        return jsonify({"ok": True, "months": months})

    @app.get(f"{API_PREFIX}/cases")
    def cr_cases():
        """既有案件列表（按规整案件名去重、拼音排序），供下拉筛选/入库合并/改案件名选择。
        跟随 ?scope=mine|dept|all 过滤：切到「本部门」时，下拉里出现的就是本部门所有人的案件。
        支持 ?month=YYYY-MM 或 ?from=...&to=... 按入库时间过滤。"""
        user = _cr_viewer()
        if user is None:
            return jsonify({"ok": False, "detail": "未登录或登录已过期"}), 401
        scope = (request.args.get("scope") or "mine").strip()
        if scope == "all" and not user.get("super_admin"):
            return jsonify({"ok": False, "detail": "无权查看全部战果"}), 403
        recs = scoped_records(user, scope)
        recs = _filter_by_period(recs,
                                 request.args.get("month") or "",
                                 request.args.get("from") or "",
                                 request.args.get("to") or "")
        recs = _filter_by_org(recs,
                              request.args.get("dept") or "",
                              request.args.get("user") or "")
        groups = {}
        for rec in recs:
            cn = parser.normalize_case_name((rec.get("fields") or {}).get("案件名"))
            if not cn:
                continue
            g = groups.setdefault(cn, {"name": "", "latest": "", "records": 0})
            g["records"] += 1
            created = rec.get("created_at") or ""
            if created >= g["latest"]:
                g["latest"] = created
                g["name"] = (rec.get("fields") or {}).get("案件名") or cn
        out = [{"name": g["name"], "records": g["records"], "normalized": cn}
               for cn, g in groups.items()]
        out.sort(key=lambda x: _case_sort_key(x["name"]))
        return jsonify({"ok": True, "count": len(out), "cases": out})

    @app.get(f"{API_PREFIX}/categories")
    def cr_categories_list():
        """既有类别集合：learned 为用户已学习的「物品名→类别」，known 为可选类别列表。"""
        learned = category_kb.all_overrides()
        builtin = set(parser.known_categories())
        used = set(learned.values())
        for rec in list_records():
            for it in rec.get("items") or []:
                if it.get("category"):
                    used.add(it["category"])
        known = sorted(builtin | used, key=lambda s: (s == "其他", s))
        return jsonify({"ok": True, "count": len(learned), "learned": learned, "known": known})

    @app.post(f"{API_PREFIX}/categories")
    def cr_categories_learn():
        """学习一条「物品名→类别」，持久化供后续解析优先采用。"""
        _set_operation("学习战果物品类别")
        body = request.get_json(silent=True) or {}
        name = str(body.get("name") or "").strip()
        category = str(body.get("category") or "").strip()
        if not name or not category:
            return jsonify({"ok": False, "detail": "需要同时提供 name 与 category"}), 400
        category_kb.set_override(name[:40], category[:20])
        return jsonify({"ok": True, "learned": category_kb.all_overrides()})

    @app.delete(f"{API_PREFIX}/categories/<item_name>")
    def cr_categories_forget(item_name):
        """删除某物品名的学习类别，恢复默认判定。"""
        _set_operation("删除战果物品类别学习")
        category_kb.remove_override(item_name)
        return jsonify({"ok": True, "learned": category_kb.all_overrides()})

    @app.get(f"{API_PREFIX}/records")
    def cr_list():
        """本地台账列表（按入库时间倒序），支持 ?case=<案件名> 仅返回该案件的记录；
        支持 ?scope=mine|dept|all（默认 mine）按范围过滤；
        支持 ?month=YYYY-MM 或 ?from=...&to=... 按入库时间过滤；
        支持 ?dept=<部门ID或名称>、?user=<用户名/姓名> 按部门/主办人过滤。"""
        user = _cr_viewer()
        if user is None:
            return jsonify({"ok": False, "detail": "未登录或登录已过期"}), 401
        scope = (request.args.get("scope") or "mine").strip()
        if scope == "all" and not user.get("super_admin"):
            return jsonify({"ok": False, "detail": "无权查看全部战果"}), 403
        case_name = parser.normalize_case_name(request.args.get("case") or "")
        recs = scoped_records(user, scope)
        recs = _filter_by_period(recs,
                                 request.args.get("month") or "",
                                 request.args.get("from") or "",
                                 request.args.get("to") or "")
        recs = _filter_by_org(recs,
                              request.args.get("dept") or "",
                              request.args.get("user") or "")
        if case_name:
            recs = [r for r in recs
                    if parser.normalize_case_name((r.get("fields") or {}).get("案件名")) == case_name]
        return jsonify({"ok": True, "count": len(recs), "records": recs})

    @app.post(f"{API_PREFIX}/records")
    def cr_create():
        """保存一条战果记录：body {fields, source_text, items?, merge_mode?, merge_case?}。

        items 可选：由前端把可编辑的物品明细（含用户修正后的类别）传回，
        未传时按缴获物品字段自动拆分。保存的同时将「物品名→类别」学习入库。

        同案检测：默认 merge_mode=auto，若规整后的案件名已存在于既有台账，
        则不落盘、返回 duplicate=true + matches（供前端提示是否合并）；
        merge_mode=merge 时将本次记录并入 matches 中指定（merge_case）的既有案件；
        merge_mode=new 时跳过同案提示、直接新增为一条新记录。
        """
        _set_operation("保存战果记录")
        body = request.get_json(silent=True) or {}
        fields = normalize_fields(body.get("fields"))
        if not any(fields.get(k) for k in FIELD_KEYS):
            return jsonify({"ok": False, "detail": "没有任何要素可保存，请先解析或补充字段"}), 400
        source_text = (body.get("source_text") or "")
        if isinstance(source_text, str) and len(source_text) > MAX_TEXT_LEN:
            source_text = source_text[:MAX_TEXT_LEN]
        items = _validate_items(body.get("items")) or []

        # 规整案件名（去引号/空白），保证同一案件不同写法在对齐与存储上一致
        if fields.get("案件名"):
            fields["案件名"] = parser.normalize_case_name(fields["案件名"])
        case_name = fields.get("案件名") or ""
        merge_mode = (body.get("merge_mode") or "auto").strip()
        if merge_mode == "merge":
            target = (body.get("merge_case") or "").strip()
            if target:
                fields["案件名"] = parser.normalize_case_name(target)
                case_name = fields["案件名"]

        user = _cr_viewer()
        if user is None:
            return jsonify({"ok": False, "detail": "未登录或登录已过期"}), 401

        # 主办大队缺省时取当前用户部门，并归一到配置可选值
        if not fields.get("主办大队") and user.get("department"):
            dept = str(user.get("department") or "").strip()
            if dept:
                fields["主办大队"] = _normalize_unit(dept)

        # 主办人为空时，以当前用户姓名为默认值
        if not fields.get("主办人"):
            fields["主办人"] = user.get("name") or ""

        # 同案检测只在自己的记录里查：合并目标必须是自己可管理的记录
        if merge_mode == "auto" and case_name:
            matches = _find_case_matches(case_name, scoped_records(user, "mine"))
            if matches:
                return jsonify({
                    "ok": True,
                    "duplicate": True,
                    "case_name": fields["案件名"],
                    "matches": matches,
                })

        with _LOCK:
            rec = {
                "id": uuid.uuid4().hex[:12],
                "fields": fields,
                "items": items,
                "source_text": source_text,
                "created_at": _now(),
                "created_by": user["username"],
                "created_by_name": user["name"],
                "unit_id": user.get("unit_id", ""),
                "department_id": user.get("department_id", ""),
            }
            save_record(rec)
        for it in items:
            if it.get("category") and it.get("name"):
                category_kb.set_override(it["name"], it["category"])
        return jsonify({"ok": True, "record": rec})

    @app.get(f"{API_PREFIX}/records/<rid>")
    def cr_get(rid):
        user = _cr_viewer()
        if user is None:
            return jsonify({"ok": False, "detail": "未登录或登录已过期"}), 401
        rec = load_record(rid)
        if not rec or not _record_readable(user, rec):
            return jsonify({"ok": False, "detail": "记录不存在或无权访问"}), 404
        return jsonify({"ok": True, "record": rec})

    @app.put(f"{API_PREFIX}/records/<rid>/case")
    def cr_update_case(rid):
        """修改某条台账记录的「案件名」：可输入新案件名，或改为既有案件的名称
        （选择既有案件时即并入该案，战果汇总「涉及 N 起」会随之按案件去重重算）。
        仅录入人本人或超级管理员可操作。"""
        _set_operation("修改战果记录案件名")
        user = _cr_viewer()
        if user is None:
            return jsonify({"ok": False, "detail": "未登录或登录已过期"}), 401
        body = request.get_json(silent=True) or {}
        new_name = parser.normalize_case_name((body.get("case_name") or "").strip())
        if not new_name:
            return jsonify({"ok": False, "detail": "案件名不能为空"}), 400
        new_name = new_name[:MAX_ITEM_LEN]
        with _LOCK:
            rec = load_record(rid)
            if not rec or not _record_readable(user, rec):
                return jsonify({"ok": False, "detail": "记录不存在或无权访问"}), 404
            if not _record_manageable(user, rec):
                return jsonify({"ok": False, "detail": "仅录入人本人或管理员可修改"}), 403
            rec["fields"] = rec.get("fields") or {}
            rec["fields"]["案件名"] = new_name
            save_record(rec)
        return jsonify({"ok": True, "record": rec})

    @app.put(f"{API_PREFIX}/records/<rid>")
    def cr_update(rid):
        """编辑台账记录：可修改 fields（五要素）/ source_text / items（物品明细）。
        仅录入人本人或超级管理员可操作；保存时同步学习「物品名→类别」。
        body：{fields?, source_text?, items?}，缺省字段保持不变。"""
        _set_operation("编辑战果记录")
        user = _cr_viewer()
        if user is None:
            return jsonify({"ok": False, "detail": "未登录或登录已过期"}), 401
        body = request.get_json(silent=True) or {}
        with _LOCK:
            rec = load_record(rid)
            if not rec or not _record_readable(user, rec):
                return jsonify({"ok": False, "detail": "记录不存在或无权访问"}), 404
            if not _record_manageable(user, rec):
                return jsonify({"ok": False, "detail": "仅录入人本人或管理员可编辑"}), 403
            # 字段：仅接收五要素，缺省保留原值
            if body.get("fields") is not None:
                fields = normalize_fields(body.get("fields"))
                if not any(fields.get(k) for k in FIELD_KEYS):
                    return jsonify({"ok": False, "detail": "没有任何要素可保存"}), 400
                if fields.get("案件名"):
                    fields["案件名"] = parser.normalize_case_name(fields["案件名"])
                # 主办人为空时默认取录入人姓名（旧记录兼容：无主办人则以录入人兜底）
                if not fields.get("主办人"):
                    fields["主办人"] = rec.get("created_by_name") or user.get("name") or ""
                rec["fields"] = fields
            # 原始报告
            if "source_text" in body:
                src = body.get("source_text") or ""
                rec["source_text"] = str(src)[:MAX_TEXT_LEN]
            # 物品明细
            if body.get("items") is not None:
                items = _validate_items(body.get("items"))
                if items is None:
                    return jsonify({"ok": False, "detail": "物品明细格式不正确"}), 400
                rec["items"] = items
            save_record(rec)
            saved = dict(rec)
        # 同步学习类别（编辑后的明细也应记住）
        for it in saved.get("items") or []:
            if it.get("category") and it.get("name"):
                category_kb.set_override(it["name"], it["category"])
        return jsonify({"ok": True, "record": saved})

    @app.delete(f"{API_PREFIX}/records/<rid>")
    def cr_delete(rid):
        _set_operation("删除战果记录")
        user = _cr_viewer()
        if user is None:
            return jsonify({"ok": False, "detail": "未登录或登录已过期"}), 401
        with _LOCK:
            rec = load_record(rid)
            if not rec or not _record_readable(user, rec):
                return jsonify({"ok": False, "detail": "记录不存在或无权访问"}), 404
            if not _record_manageable(user, rec):
                return jsonify({"ok": False, "detail": "仅录入人本人或管理员可删除"}), 403
            try:
                os.remove(_record_path(rid))
            except OSError:
                return jsonify({"ok": False, "detail": "删除失败"}), 500
        return jsonify({"ok": True})

    @app.get(f"{API_PREFIX}/export")
    def cr_export():
        """按当前筛选条件导出战果台账为 Excel（.xlsx）。

        过滤参数与 /records 一致：scope / case / month / from / to / dept / user。
        导出的记录包含：案件名 / 时间 / 主办大队 / 主办人 / 抓获人数 / 缴获物品 /
        缴获明细 / 录入人 / 所属部门 / 入库时间 / 原始报告。
        """
        _set_operation("导出战果台账")
        user = _cr_viewer()
        if user is None:
            return jsonify({"ok": False, "detail": "未登录或登录已过期"}), 401
        scope = (request.args.get("scope") or "mine").strip()
        if scope == "all" and not user.get("super_admin"):
            return jsonify({"ok": False, "detail": "无权查看全部战果"}), 403
        if not OPENPYXL_AVAILABLE:
            return jsonify({"ok": False, "detail": "缺少 openpyxl，无法导出 Excel（pip install openpyxl）"}), 500
        case_name = parser.normalize_case_name(request.args.get("case") or "")
        recs = scoped_records(user, scope)
        recs = _filter_by_period(recs,
                                 request.args.get("month") or "",
                                 request.args.get("from") or "",
                                 request.args.get("to") or "")
        recs = _filter_by_org(recs,
                              request.args.get("dept") or "",
                              request.args.get("user") or "")
        if case_name:
            recs = [r for r in recs
                    if parser.normalize_case_name((r.get("fields") or {}).get("案件名")) == case_name]
        recs.sort(key=lambda r: r.get("created_at") or "", reverse=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "战果台账"
        headers = ["案件名", "时间", "主办大队", "主办人", "抓获人数", "涉案价值",
                   "缴获物品", "缴获明细", "录入人", "所属部门", "入库时间", "原始报告"]
        ws.append(headers)
        # 表头样式
        header_fill = PatternFill("solid", fgColor="4A90D9")
        thin = Side(style="thin", color="BBBBBB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        # 数据行
        for rec in recs:
            f = rec.get("fields") or {}
            items_txt = "；".join(
                f"{it.get('name') or ''}"
                f"{('×' + str(it.get('quantity'))) if it.get('quantity') is not None else ''}"
                f"{(it.get('unit') or '')}"
                for it in (rec.get("items") or [])
            )
            dept_name = rec.get("department_name") or rec.get("department_id") or ""
            ws.append([
                f.get("案件名", ""),
                f.get("时间", ""),
                f.get("主办大队", ""),
                f.get("主办人", "") or rec.get("created_by_name") or "",
                f.get("抓获人数", ""),
                f.get("涉案价值", ""),
                f.get("缴获物品", ""),
                items_txt,
                rec.get("created_by_name") or rec.get("created_by") or "",
                dept_name,
                rec.get("created_at", ""),
                rec.get("source_text", "") or "",
            ])
        # 列宽与自动换行
        widths = [24, 16, 16, 12, 12, 16, 30, 30, 12, 16, 20, 40]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"case-report-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.get(f"{API_PREFIX}/records/<rid>/download")
    def cr_download(rid):
        """下载单条记录的键值对 JSON 文件。"""
        _set_operation("下载战果记录")
        user = _cr_viewer()
        if user is None:
            return jsonify({"ok": False, "detail": "未登录或登录已过期"}), 401
        rec = load_record(rid)
        if not rec or not _record_readable(user, rec):
            return jsonify({"ok": False, "detail": "记录不存在或无权访问"}), 404
        data = json.dumps(rec, ensure_ascii=False, indent=2).encode("utf-8")
        buf = io.BytesIO(data)
        return send_file(buf, as_attachment=True,
                         download_name=f"case-{rid}.json",
                         mimetype="application/json")