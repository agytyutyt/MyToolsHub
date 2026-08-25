"""轨迹转换 —— JZToolsHub 后端插件路由。

功能：读取 Excel 轨迹表（至少含时间 / 经度 / 纬度三列，列名不统一可通过
插件目录下 config.json 自定义，默认「开始时间 / 经度 / 纬度」），从最早时间
起按固定时间间隔抽样，若目标时刻无数据则取距离该时刻最近的一条（保留原始
时间 / 经度 / 纬度），抽样结果封装为 JSON，再经 QR-transfer 协议编码为
二维码视频流（zfec 前向纠错 + qrcode 渲染 + opencv 写视频）。

协议细节与 https://github.com/Zhen-Ni/QR-transfer 的 qrtransfer.py 保持一致：
- 每帧 QR 内容为 base64 文本：前 4 字符=帧序号、4-8 字符=总帧数、
  8-12 字符=纠错参数(k,m)、12 字符之后=该帧数据分块(base64)；
- 原始数据带 4 字节长度前缀，经 zfec 分块/纠错后逐帧入码。

并发设计（对齐 qr-video-decode 插件）：
- POST /convert 接收 Excel 与参数，立即返回 task_id；
- 解析、抽样、编码、写视频放到后台线程执行，按「帧」粒度更新进度；
- 前端通过 GET /status/<task_id> 轮询，完成后经 /download/<task_id> 取视频。
"""

import base64
import io
import json
import math
import os
import re
import threading
import time
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta

from flask import jsonify, request, send_file

try:
    from jztools_admin.routes import get_session_user
except Exception:  # admin 插件缺失时兜底（理论上不会发生）
    get_session_user = None

API_PREFIX = "/api/trajectory-convert"

# ---- 基础量（与 QR-transfer 参考实现一致） ----
MAX_FEC_M = 256
SIZE_INDEX = 3          # base64 编码前的头部各字段字节数
SIZE_PREFIX = 4         # base64 字符数（3 字节 → 4 字符）
SIZE_DATASIZE = 4       # 原始数据长度前缀字节数
BYTE_ORDER = "big"
BOX_SIZE = 8             # 二维码渲染模块边长（像素）；值越大视频越清晰、边界越锐利
FRAMERATE = 15
# 视频编码：使用 avc1(H.264) 保证浏览器 <video> 可直接播放（mp4v 浏览器不支持，
# 导致 QR 视频流解码插件无法逐帧读取）。openpyxl 环境下 opencv 内置 FFmpeg 支持 avc1。
VIDEO_FOURCC = "avc1"
FEC_RATIO = 0.1         # 前向纠错比例：额外生成 1/(1-fec_ratio) 帧
DEFAULT_FIELDS = {
    "time_field": "开始时间",
    "lng_field": "经度",
    "lat_field": "纬度",
}

# 时间字符串常见格式
TIME_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y/%m/%d %H:%M:%S",
    "%Y/%m/%d %H:%M",
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y/%m/%d %H:%M:%S.%f",
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%Y%m%d%H%M%S",
    "%Y年%m月%d日 %H:%M:%S",
    "%Y年%m月%d日%H:%M:%S",
    "%Y.%m.%d %H:%M:%S",
)

# ---- 依赖可用性探测（缺失时给出明确安装提示） ----
try:
    import qrcode
    from qrcode import constants as qrcode_constants
    from qrcode.exceptions import DataOverflowError as QR_OVERFLOW
    QRCODE_AVAILABLE = True
except Exception:  # pragma: no cover
    qrcode = None
    QRCODE_AVAILABLE = False

try:
    import zfec
    ZFEC_AVAILABLE = True
except Exception:  # pragma: no cover
    zfec = None
    ZFEC_AVAILABLE = False

try:
    import cv2
    CV2_AVAILABLE = True
except Exception:  # pragma: no cover
    cv2 = None
    CV2_AVAILABLE = False

try:
    import numpy as np
    NUMPY_AVAILABLE = True
except Exception:  # pragma: no cover
    np = None
    NUMPY_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except Exception:  # pragma: no cover
    xlrd = None
    XLRD_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except Exception:  # pragma: no cover
    openpyxl = None
    OPENPYXL_AVAILABLE = False

# 后台任务
CONVERT_WORKERS = 2
_executor = ThreadPoolExecutor(max_workers=CONVERT_WORKERS)
TASKS = {}
TASKS_LOCK = threading.Lock()
TASK_TTL_SECONDS = 30 * 60

_BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
_CONFIG_PATH = os.path.join(_BACKEND_DIR, "config.json")
_TASK_DIR = os.path.join(_BACKEND_DIR, ".task_cache")


# ===================== 配置读取 =====================

def load_config():
    """读取插件目录 config.json 中的字段名默认值。"""
    cfg = dict(DEFAULT_FIELDS)
    try:
        if os.path.isfile(_CONFIG_PATH):
            # utf-8-sig 兼容带/不带 BOM 的配置文件
            with open(_CONFIG_PATH, "r", encoding="utf-8-sig") as f:
                cfg.update(json.load(f))
    except Exception:
        pass
    return cfg


# ===================== 任务状态 =====================

def set_task(task_id, **kwargs):
    with TASKS_LOCK:
        TASKS[task_id] = {**TASKS.get(task_id, {}), **kwargs}


def get_task(task_id):
    with TASKS_LOCK:
        return TASKS.get(task_id)


def cleanup_tasks():
    now = time.time()
    with TASKS_LOCK:
        expired = [
            tid for tid, t in TASKS.items()
            if t.get("created_at", 0) < now - TASK_TTL_SECONDS
        ]
        for tid in expired:
            TASKS.pop(tid, None)


def _task_file(task_id, ext=".mp4"):
    return os.path.join(_TASK_DIR, f"{task_id}{ext}")


def _tc_viewer():
    """当前登录用户上下文；未登录返回 None。"""
    if get_session_user is None:
        return None
    try:
        return get_session_user()
    except Exception:
        return None


def _task_owned_by(task, user):
    """任务归属校验：任务元数据记录过创建者时，非创建者（且非超管）视为不存在（404）。

    任务 ID 为随机 UUID + TTL 定期清理，本身已难猜测；此校验是纵深防御（R6）。
    """
    owner = task.get("created_by")
    if not owner:
        return True
    if user is None:
        return False
    if user.get("super_admin"):
        return True
    return owner == user.get("username")


def _static_image_file(task_id, index):
    """第 index（1 起）张静态二维码图片的缓存路径。"""
    return os.path.join(_TASK_DIR, f"{task_id}_{index:02d}.png")


def _clean_task_files():
    os.makedirs(_TASK_DIR, exist_ok=True)
    now = time.time()
    for name in os.listdir(_TASK_DIR):
        path = os.path.join(_TASK_DIR, name)
        try:
            if now - os.path.getmtime(path) > TASK_TTL_SECONDS:
                os.remove(path)
        except Exception:
            pass


# ===================== Excel 解析 =====================

def _read_excel_rows(path, filename):
    """读取 Excel 全部单元格（含表头），按行返回，.xlsx/.xls 自适应。"""
    ext = (filename or path).lower().rsplit(".", 1)[-1]
    if ext == "xlsx":
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("后端缺少 openpyxl，无法解析 .xlsx，请执行：pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
        return rows
    if ext == "xls":
        if not XLRD_AVAILABLE:
            raise RuntimeError("后端缺少 xlrd，无法解析 .xls，请执行：pip install xlrd")
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        rows = []
        for r in range(sheet.nrows):
            row = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    row.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                else:
                    row.append(cell.value)
            rows.append(row)
        return rows
    raise ValueError(f"不支持的 Excel 格式：.{ext}（仅支持 .xlsx / .xls）")


def _clean_text(v):
    if v is None:
        return ""
    if isinstance(v, str):
        # 归一化全角字符与常见空白
        text = v.replace("\u3000", " ").replace("（", "(").replace("）", ")")
        text = text.replace("，", ",")
        return text.strip().strip("\uFEFF")
    return v


def _find_header(rows, fields):
    """在前若干行中定位表头所在行，返回 (header_index, col_map)。

    col_map: {字段名 -> 列下标}；找不到对应列时抛出带可选项的提示。
    """
    names = list(fields.values())
    for i, row in enumerate(rows[:6]):
        cols = {str(_clean_text(c) or ""): j for j, c in enumerate(row)}
        found = {n: cols.get(n) for n in names}
        if found[names[0]] is not None and found[names[1]] is not None and found[names[2]] is not None:
            return i, found
    # 找不到完整表头：给出最接近的列名提示
    first = rows[0] if rows else []
    headers = [str(_clean_text(c) or "") for c in first]
    hint = "、".join(f"「{h}」" for h in headers[:10]) or "无"
    raise ValueError(
        f"未能在表格表头中找到字段：{fields['time_field']} / {fields['lng_field']} / "
        f"{fields['lat_field']}。当前表头：{hint}（可在插件页面或 backend/config.json 自定义字段名）"
    )


def _parse_time(cell):
    """把单元格的值解析为 datetime，失败返回 None。"""
    if isinstance(cell, datetime):
        return cell
    c = _clean_text(cell)
    if isinstance(c, (int, float)):
        if c == 0 and cell == 0:
            return None
        # 常用时间尺度识别
        try:
            f = float(c)
        except (TypeError, ValueError):
            return None
        if 21915 <= f <= 60000:  # 1960-2064，视为 Excel 序列日期（1899-12-30 起）
            return datetime(1899, 12, 30) + timedelta(days=f)
        if f > 10 ** 12:  # 毫秒时间戳
            return datetime.fromtimestamp(f / 1000.0)
        if f > 10 ** 9:  # 秒时间戳
            return datetime.fromtimestamp(f)
        return None
    if isinstance(c, str):
        s = c.strip()
        for fmt in TIME_FORMATS:
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


def _format_time(cell, dt):
    """保留原始时间的可读字符串：字符串原样，日期单元格统一格式化。"""
    if isinstance(_clean_text(cell), str):
        s = str(_clean_text(cell))
        if s:
            return s
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _to_float(cell, label):
    """把经纬度单元格转为 float，支持 DMS(度分秒) 与简单清洗。"""
    if isinstance(cell, (int, float)):
        try:
            return float(cell)
        except (TypeError, ValueError):
            pass
    c = _clean_text(cell)
    if not isinstance(c, str):
        raise ValueError(f"{label} 无法识别为数值")
    s = c.replace("\u00b0", " ").replace("°", " ").replace("′", "'").replace("′", "'")
    s = s.replace("″", '"').replace("〃", '"')
    m = re.match(r"^\s*([+-]?[\d.]+)(?:\s*([\d.]+))?(?:\s*([\d.]+))?\s*[NWES]?\s*$", s, re.IGNORECASE)
    if m:
        d, mi, sec = (float(m.group(i)) if m.group(i) else 0.0 for i in (1, 2, 3))
        v = d + mi / 60.0 + sec / 3600.0
        if s.strip() and s.strip()[-1].upper() in "SW":
            v = -abs(v)
        return v
    raise ValueError(f"{label} 无法识别为数值：{c}")


def parse_trajectory(rows, cfg):
    """从表头+数据行解析轨迹，返回按时间升序的 [(dt, lng, lat, time_text)]。"""
    header_idx, col_map = _find_header(rows, cfg)
    t_i, lng_i, lat_i = (col_map[cfg["time_field"]], col_map[cfg["lng_field"]], col_map[cfg["lat_field"]])

    points = []
    for row in rows[header_idx + 1:]:
        t_cell = row[t_i] if t_i < len(row) else None
        lng_cell = row[lng_i] if lng_i < len(row) else None
        lat_cell = row[lat_i] if lat_i < len(row) else None
        if _clean_text(t_cell) == "" or _clean_text(lng_cell) == "" or _clean_text(lat_cell) == "":
            continue
        dt = _parse_time(t_cell)
        if dt is None:
            continue
        try:
            lng = _to_float(lng_cell, cfg["lng_field"])
            lat = _to_float(lat_cell, cfg["lat_field"])
        except ValueError:
            continue
        # 经纬度均为 0 视为脏数据，排除
        if lng == 0 and lat == 0:
            continue
        points.append((dt, lng, lat, _format_time(t_cell, dt)))
    if not points:
        raise ValueError(
            f"未能解析到有效轨迹数据（需要 {cfg['time_field']} / {cfg['lng_field']} / "
            f"{cfg['lat_field']} 三列的有效行）"
        )
    points.sort(key=lambda p: p[0])
    return points


# ===================== 脏数据过滤 =====================

EARTH_RADIUS_KM = 6371.0   # 地球平均半径（公里）
MAX_SPEED_KMH = 120.0      # 移动速度阈值，超过视为脏数据


def haversine_km(lng1, lat1, lng2, lat2):
    """通过经纬度计算两点间的球面（大圆）距离，单位公里。"""
    lat1, lng1, lat2, lng2 = map(math.radians, (lat1, lng1, lat2, lng2))
    dlat = lat2 - lat1
    dlng = lng2 - lng1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlng / 2) ** 2
    return 2 * EARTH_RADIUS_KM * math.asin(math.sqrt(a))


def is_excessive_speed(prev, cur):
    """判断两点间移动速度是否超过 120km/h（结合距离与时间差），视为脏数据。

    prev/cur 为 (dt, lng, lat, time_text)。时间差不大于 0（无法计算速度）
    或距离过小（<1m，坐标噪声抖动）时不判脏。
    """
    hours = (cur[0] - prev[0]).total_seconds() / 3600.0
    if hours <= 0:
        return False
    dist = haversine_km(prev[1], prev[2], cur[1], cur[2])
    if dist < 0.001:
        return False
    return dist / hours > MAX_SPEED_KMH


# ===================== 时间抽样 =====================

def sample_points(points, interval_minutes):
    """从最早时间起每隔 interval 分钟取一条：取距离目标时刻最近的数据。

    末尾目标时刻推进到覆盖末条记录的整倍间隔，以保证末条记录附近必有采样点。
    取点过程中结合时间差计算相邻两点的距离与移动速度，速度超过 120km/h 的
    点位视为脏数据，排除（首点始终保留）。
    返回 (sampled, n_rows)。sampled 元素为原始行。
    """
    times = [p[0] for p in points]
    start, end = times[0], times[-1]
    span = (end - start).total_seconds() / 60.0
    n_steps = int(math.ceil(span / interval_minutes)) if interval_minutes else 0
    targets = [start + timedelta(minutes=i * interval_minutes) for i in range(n_steps + 1)]

    def nearest(target):
        lo, hi = 0, len(times) - 1
        while lo < hi:
            mid = (lo + hi) // 2
            if times[mid] < target:
                lo = mid + 1
            else:
                hi = mid
        # lo 是第一个 >= target 的下标；比较 lo-1 与其本身
        cand = [lo - 1, lo]
        best = None
        for idx in cand:
            if 0 <= idx < len(times):
                if best is None or abs(times[idx] - target) < abs(times[best] - target):
                    best = idx
        return points[best]

    sampled = []
    last = None
    for t in targets:
        pt = nearest(t)
        # 与上一个保留点比较：移动速度超过 120km/h 视为脏数据，排除
        if last is not None and is_excessive_speed(last, pt):
            continue
        sampled.append(pt)
        last = pt
    return sampled, len(times)


# ===================== JSON 封装 =====================

def wrap_json(sampled, cfg, interval_minutes):
    """封装为纯键值对 JSON：键为时间（保留原始时间），值为 [经度, 纬度]。

    不含其他元数据（type / interval / count / fields 等均不写入）。
    """
    return {p[3]: [p[1], p[2]] for p in sampled}


# ===================== QR-transfer 编码 =====================

def encode_index(n):
    b = n.to_bytes(SIZE_INDEX, BYTE_ORDER)
    return base64.encodebytes(b).replace(b"\n", b"")


def encode_fecinfo(k, m):
    return encode_index((k << 8) + m)


def encode_data(b):
    return base64.encodebytes(b).replace(b"\n", b"")


def _chunk_data(data, version, err):
    """按 QR 版本容量切块（对齐 qrtransfer.py 的 _chunk_data）。"""
    maxchunksize = qrcode.util.BIT_LIMIT_TABLE[err][version]
    chunk_prefix_size = max(qrcode.util.mode_sizes_for_version(version).values()) + 4
    maxchunksize = (maxchunksize - chunk_prefix_size) // 8
    maxchunksize -= SIZE_PREFIX * 3
    maxchunksize = int(maxchunksize / 4) * 3
    if maxchunksize <= 0:
        raise ValueError(f"QR 版本 {version} 过小，无法承载数据（请调大二维码版本）")
    nchunks = max(1, math.ceil(len(data) / maxchunksize))
    chunks = []
    for i in range(nchunks):
        beg = i * maxchunksize
        chunk = data[beg:beg + maxchunksize]
        if len(chunk) < maxchunksize:
            chunk += b"\0" * (maxchunksize - len(chunk))
        chunks.append(chunk)
    return chunks


def _fec_encode(data_list, fec_ratio=FEC_RATIO):
    """zfec 前向纠错，返回 (share_list, k, m)。"""
    size = len(data_list)
    chunksize = len(data_list[0])
    blocksize = math.ceil(size / (1 - fec_ratio))
    nblocks = math.ceil(blocksize / MAX_FEC_M)
    blocksize = math.ceil(blocksize / nblocks)
    if size % nblocks:
        data_list = data_list + [b"\0" * chunksize] * (nblocks - size % nblocks)
    m = blocksize
    k = len(data_list) // nblocks
    encoder = zfec.Encoder(k, m)
    mapped = []
    for i in range(nblocks):
        mapped.append(encoder.encode(data_list[i::nblocks]))
    out = []
    for i in range(blocksize):
        for j in range(nblocks):
            out.append(mapped[j][i])
    return out, k, m


def _qrformat_encode(share_list, k, m):
    nframes = len(share_list)
    codes = []
    for i, share in enumerate(share_list):
        header = encode_index(i) + encode_index(nframes) + encode_fecinfo(k, m)
        codes.append((header + encode_data(share)).decode("ascii"))
    return codes


def encode_to_video(raw_bytes, version, out_path, progress_cb=None):
    """把字节数据按 QR-transfer 编码为二维码视频文件。返回 (nframes, k, m)。"""
    if not (QRCODE_AVAILABLE and ZFEC_AVAILABLE and CV2_AVAILABLE and NUMPY_AVAILABLE):
        missing = [
            name for name, ok in (
                ("qrcode", QRCODE_AVAILABLE), ("zfec", ZFEC_AVAILABLE),
                ("opencv-python(cv2)", CV2_AVAILABLE), ("numpy", NUMPY_AVAILABLE),
            ) if not ok
        ]
        raise RuntimeError("后端缺少依赖：" + "、".join(missing) + "，请先安装：pip install " + " ".join(
            n for n in ("qrcode", "zfec", "opencv-python", "numpy")
        ))

    err = qrcode_constants.ERROR_CORRECT_L
    payload = len(raw_bytes).to_bytes(SIZE_DATASIZE, BYTE_ORDER) + raw_bytes
    chunks = _chunk_data(payload, version, err)
    share_list, k, m = _fec_encode(chunks)
    codes = _qrformat_encode(share_list, k, m)

    total = len(codes)
    writer = None
    try:
        for i, code in enumerate(codes):
            qr = qrcode.QRCode(version=version, error_correction=err, box_size=BOX_SIZE)
            qr.add_data(code, optimize=0)
            qr.make(fit=False)
            gray = np.array(qr.make_image(fill_color="black", back_color="white")
                            .convert("L"), dtype=np.uint8)
            img = cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)  # FFMPEG 要求 3 通道
            if writer is None:
                h, w = img.shape[:2]
                writer = cv2.VideoWriter(out_path, cv2.VideoWriter_fourcc(*VIDEO_FOURCC),
                                         FRAMERATE, (w, h))
            writer.write(img)
            if progress_cb and (i % 4 == 0 or i == total - 1):
                progress_cb(i + 1, total, codes[i])
    finally:
        if writer is not None:
            writer.release()
    return total, k, m


def _render_qr_image(data_bytes, version, err, out_path):
    """按固定版本把字节数据渲染为一张二维码 PNG 图片。"""
    qr = qrcode.QRCode(version=version, error_correction=err, box_size=BOX_SIZE, border=4)
    qr.add_data(data_bytes, optimize=0)
    qr.make(fit=False)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(out_path)
    return out_path


def _static_json_chunks(payload, version, err):
    """把键值对数据拆分为多份可直接读取的 JSON（每份单张二维码可容纳）。

    按点位顺序累加，每份编码后的字节数不超该二维码版本的容量。
    """
    try:
        capacity = qrcode.util.BIT_LIMIT_TABLE[err][version] // 8
    except KeyError:
        capacity = 800
    budget = max(48, capacity - 16)   # 预留花括号等结构开销与安全余量
    chunks, cur = [], {}
    cur_size = 2   # "{}"
    for k, v in payload.items():
        item = json.dumps({k: v}, ensure_ascii=False, separators=(",", ":"))
        item_bytes = ("," + item[1:-1]).encode("utf-8")
        if cur and cur_size + len(item_bytes) > budget:
            chunks.append(cur)
            cur, cur_size = {}, 2
        cur[k] = v
        cur_size += len(item_bytes)
    if cur:
        chunks.append(cur)
    return chunks


def encode_static_output(payload, version, out_dir, prefix):
    """生成静态二维码图片文件，返回文件路径列表。

    数据能放进单张时输出 1 张：内容为整份 JSON，任何扫码器均可直接读取，
    也可直接导入「地图标点」模块识别出全部轨迹点；
    放不下时按点位顺序拆分为多张，每张都是独立可读的 JSON 子集
    （可直接导入「地图标点」识别其中的点位）。
    文件命名：{prefix}_01.png（单张）、{prefix}_01.png … {prefix}_NN.png（多张）。
    """
    if not QRCODE_AVAILABLE:
        raise RuntimeError("后端缺少 qrcode 依赖，请先安装：pip install qrcode")
    err = qrcode_constants.ERROR_CORRECT_L
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    single = os.path.join(out_dir, f"{prefix}_01.png")
    try:
        _render_qr_image(raw, version, err, single)
        return [single]
    except QR_OVERFLOW:
        pass
    chunks = _static_json_chunks(payload, version, err)
    if len(chunks) <= 1:
        raise ValueError("数据无法拆分为多张静态二维码，请调大二维码版本或改用「二维码视频流」模式")
    paths = []
    for i, chunk in enumerate(chunks):
        b = json.dumps(chunk, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
        p = os.path.join(out_dir, f"{prefix}_{i + 1:02d}.png")
        _render_qr_image(b, version, err, p)
        paths.append(p)
    return paths


# ===================== 后台转换任务 =====================

def _run_convert(task_id, filename, file_bytes, interval_minutes, qr_version, cfg, mode="video"):
    try:
        set_task(task_id, status="running", created_at=time.time())
        tmp_xlsx = os.path.join(_TASK_DIR, f"{task_id}_in{os.path.splitext(filename)[1] or '.xlsx'}")
        os.makedirs(_TASK_DIR, exist_ok=True)
        with open(tmp_xlsx, "wb") as f:
            f.write(file_bytes)

        set_task(task_id, progress=0.02, stage="解析 Excel")
        rows = _read_excel_rows(tmp_xlsx, filename)
        points = parse_trajectory(rows, cfg)
        set_task(task_id, progress=0.12, stage="时间抽样",
                 rows=len(points))

        sampled, total_rows = sample_points(points, interval_minutes)
        payload = wrap_json(sampled, cfg, interval_minutes)
        payload_bytes = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        base_name = os.path.splitext(filename)[0] or "轨迹"

        if mode == "static":
            # 生成静态二维码：单张能装下则一张，装不下自动拆分为多张
            img_paths = encode_static_output(payload, qr_version, _TASK_DIR, task_id)
            count = len(img_paths)
            if count == 1:
                single = _task_file(task_id, ".png")
                os.replace(img_paths[0], single)
                size = os.path.getsize(single) if os.path.exists(single) else 0
                set_task(task_id, status="done", progress=1.0,
                         output_type="static", rows=len(points),
                         base_name=base_name,
                         image_count=1,
                         image_size=size,
                         image_name=f"{base_name}二维码.png",
                         data_json=json.dumps(payload, ensure_ascii=False))
            else:
                zip_path = _task_file(task_id, ".zip")
                with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                    for i, p in enumerate(img_paths):
                        zf.write(p, f"{base_name}_第{i + 1}张（共{count}张）.png")
                size = os.path.getsize(zip_path) if os.path.exists(zip_path) else 0
                set_task(task_id, status="done", progress=1.0,
                         output_type="static", rows=len(points),
                         base_name=base_name,
                         image_count=count,
                         zip_size=size,
                         zip_name=f"{base_name}二维码（共{count}张）.zip",
                         data_json=json.dumps(payload, ensure_ascii=False))
            return

        def progress(done, total, _code):
            # 解析+抽样约 12%，编码视频 12%→100%
            set_task(task_id, progress=round(0.12 + 0.88 * done / total, 4),
                     stage=f"QR 编码 {done}/{total} 帧")

        out_path = _task_file(task_id, ".mp4")
        nframes, k, m = encode_to_video(payload_bytes, qr_version, out_path,
                                        progress_cb=progress)
        size = os.path.getsize(out_path) if os.path.exists(out_path) else 0

        set_task(task_id, status="done", progress=1.0,
                 output_type="video", rows=len(points),
                 nframes=nframes, k=k, m=m,
                 video_size=size,
                 video_name=f"{base_name}二维码流.mp4",
                 data_json=json.dumps(payload, ensure_ascii=False))
    except ValueError as e:
        set_task(task_id, status="error", detail=str(e))
    except RuntimeError as e:
        set_task(task_id, status="error", detail=str(e))
    except Exception as e:
        set_task(task_id, status="error", detail=f"转换失败：{e}")
    finally:
        try:
            os.remove(tmp_xlsx)
        except Exception:
            pass


# ===================== 路由注册 =====================

def register(app) -> None:
    """插件入口：由 JZToolsHub 主应用在启动时调用。"""

    @app.get(f"{API_PREFIX}/config")
    def tc_config():
        return jsonify({"ok": True, **load_config()})

    @app.get(f"{API_PREFIX}/status")
    def tc_status():
        return jsonify({
            "ok": True,
            "openpyxl": OPENPYXL_AVAILABLE,
            "xlrd": XLRD_AVAILABLE,
            "qrcode": QRCODE_AVAILABLE,
            "zfec": ZFEC_AVAILABLE,
            "cv2": CV2_AVAILABLE,
            "numpy": NUMPY_AVAILABLE,
        })

    @app.post(f"{API_PREFIX}/convert")
    def tc_convert():
        """接收 Excel 与抽样参数，后台执行并返回 task_id。"""
        f = request.files.get("file")
        if not f or not f.filename:
            return jsonify({"ok": False, "detail": "请选择 Excel 文件"}), 400
        if f.filename.lower().rsplit(".", 1)[-1] not in ("xlsx", "xls"):
            return jsonify({"ok": False, "detail": "仅支持 .xlsx / .xls 文件"}), 400

        # 字段名仅从后端配置文件读取（plugins/trajectory-convert/backend/config.json），
        # 不在页面展示，也不接受前端覆盖。
        cfg = load_config()

        try:
            interval = int(float(request.form.get("interval_minutes", 10)))
        except (TypeError, ValueError):
            interval = 10
        if interval <= 0:
            return jsonify({"ok": False, "detail": "时间间隔必须为正整数（分钟）"}), 400

        try:
            version = int(float(request.form.get("qr_version", 15)))
        except (TypeError, ValueError):
            version = 15
        if not 1 <= version <= 40:
            return jsonify({"ok": False, "detail": "二维码版本须在 1-40 之间"}), 400

        mode = request.form.get("mode", "video")
        if mode not in ("video", "static"):
            mode = "video"

        task_id = uuid.uuid4().hex[:12]
        cleanup_tasks()
        _clean_task_files()
        viewer = _tc_viewer()
        set_task(task_id, status="pending", progress=0.0, created_at=time.time(),
                 created_by=(viewer or {}).get("username", ""))
        filename = os.path.basename(f.filename or "trajectory.xlsx")
        file_bytes = f.read()
        _executor.submit(_run_convert, task_id, filename, file_bytes, interval, version, cfg, mode)
        return jsonify({"ok": True, "task_id": task_id})

    @app.get(f"{API_PREFIX}/status/<task_id>")
    def tc_convert_status(task_id):
        task = get_task(task_id)
        if not task:
            return jsonify({"ok": False, "detail": "任务不存在或已过期"}), 404
        if not _task_owned_by(task, _tc_viewer()):
            return jsonify({"ok": False, "detail": "任务不存在或已过期"}), 404
        payload = {
            "ok": True,
            "status": task.get("status"),
            "progress": task.get("progress", 0.0),
            "stage": task.get("stage", "") or "",
        }
        if task.get("status") == "done":
            payload.update({
                "output_type": task.get("output_type", "video"),
                "rows": task.get("rows", 0),
                "nframes": task.get("nframes", 0),
                "k": task.get("k"),
                "m": task.get("m"),
                "video_size": task.get("video_size", 0),
                "image_count": task.get("image_count", 0),
                "image_size": task.get("image_size", 0),
                "zip_size": task.get("zip_size", 0),
                "video_name": task.get("video_name", "trajectory.mp4"),
                "image_name": task.get("image_name", "trajectory.png"),
                "zip_name": task.get("zip_name", "trajectory.zip"),
                "data_json": task.get("data_json", ""),
            })
        if task.get("status") == "error":
            payload["detail"] = task.get("detail", "转换失败")
        return jsonify(payload)

    @app.get(f"{API_PREFIX}/download/<task_id>")
    @app.get(f"{API_PREFIX}/download/<task_id>/<path:filename>")
    def tc_download(task_id, filename=None):
        """返回生成的二维码文件（视频 / 单张静态图片 / 多张静态图片 ZIP）。

        路由带 filename 时 URL 以 .mp4/.png/.zip 结尾，浏览器右键另存为时能取到
        与 Content-Disposition 一致的文件名（解决「名称不一致、无后缀」问题）。
        filename 仅用于美化 URL，实际文件名以任务记录为准。
        """
        task = get_task(task_id)
        if not task or task.get("status") != "done":
            return jsonify({"ok": False, "detail": "任务不存在或未完成"}), 404
        if not _task_owned_by(task, _tc_viewer()):
            return jsonify({"ok": False, "detail": "任务不存在或未完成"}), 404
        if task.get("output_type") == "static":
            if task.get("image_count", 1) > 1:
                # 多张静态二维码：打包为 ZIP 下载
                path = _task_file(task_id, ".zip")
                if not os.path.isfile(path):
                    return jsonify({"ok": False, "detail": "ZIP 文件已被清理"}), 404
                return send_file(
                    path,
                    mimetype="application/zip",
                    as_attachment=True,
                    download_name=task.get("zip_name", "trajectory.zip"),
                )
            path = _task_file(task_id, ".png")
            if not os.path.isfile(path):
                return jsonify({"ok": False, "detail": "图片文件已被清理"}), 404
            return send_file(
                path,
                mimetype="image/png",
                as_attachment=True,
                download_name=task.get("image_name", "trajectory.png"),
            )
        path = _task_file(task_id, ".mp4")
        if not os.path.isfile(path):
            return jsonify({"ok": False, "detail": "视频文件已被清理"}), 404
        return send_file(
            path,
            mimetype="video/mp4",
            as_attachment=True,
            download_name=task.get("video_name", "trajectory.mp4"),
        )

    @app.get(f"{API_PREFIX}/image/<task_id>/<int:index>")
    def tc_static_image(task_id, index):
        """返回某一张静态二维码图片（供页面预览）。index 从 1 开始。"""
        task = get_task(task_id)
        if (not task or task.get("status") != "done"
                or task.get("output_type") != "static"):
            return jsonify({"ok": False, "detail": "任务不存在或未完成"}), 404
        if not _task_owned_by(task, _tc_viewer()):
            return jsonify({"ok": False, "detail": "任务不存在或未完成"}), 404
        total = task.get("image_count", 1)
        if not 1 <= index <= total:
            return jsonify({"ok": False, "detail": "图片序号非法"}), 404
        path = _static_image_file(task_id, index)
        if not os.path.isfile(path):
            return jsonify({"ok": False, "detail": "图片已被清理"}), 404
        return send_file(path, mimetype="image/png")

    @app.post(f"{API_PREFIX}/download-selected")
    def tc_download_selected():
        """把用户勾选的多张静态二维码打包为 ZIP 返回。

        请求体：{"task_id": "...", "indices": [1, 3, 5]}
        """
        data = request.get_json(silent=True) or {}
        task_id = data.get("task_id")
        indices = data.get("indices")
        task = get_task(task_id)
        if (not task or task.get("status") != "done"
                or task.get("output_type") != "static"):
            return jsonify({"ok": False, "detail": "任务不存在或未完成"}), 404
        if not _task_owned_by(task, _tc_viewer()):
            return jsonify({"ok": False, "detail": "任务不存在或未完成"}), 404
        total = task.get("image_count", 1)
        try:
            idxs = sorted({int(i) for i in (indices or [])})
        except (TypeError, ValueError):
            return jsonify({"ok": False, "detail": "图片序号非法"}), 400
        idxs = [i for i in idxs if 1 <= i <= total]
        if not idxs:
            return jsonify({"ok": False, "detail": "未选择图片"}), 400

        base = task.get("base_name") or "轨迹"
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for i in idxs:
                p = _static_image_file(task_id, i)
                if os.path.isfile(p):
                    zf.write(p, f"{base}_第{i}张（共{len(idxs)}张）.png")
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/zip",
            as_attachment=True,
            download_name=f"{base}二维码（所选{len(idxs)}张）.zip",
        )
